#!/usr/bin/env python3
"""
Merge the three idea pools (M1 Competitor, M3 Reddit, M2 Other-niche) into one
clubbed-ideas.csv, per the agreed schema. Stacked M1 -> M3 -> M2, each block
keeping its own sort order (blocks are contiguous by row order — CSV has no colour).
All RAG / candidate-content / verdict columns are created here but left EMPTY;
they are filled by the reuse check (rag.py retrieve, then the Stage-3 LLM pass).

Run: build_clubbed.py <asset-engine dir>
"""
import csv, os, sys
import openpyxl
csv.field_size_limit(1 << 24)

AE = os.path.abspath(sys.argv[1]) if len(sys.argv) > 1 else "."
M1 = os.path.join(AE, "competitor-study", "competitor-study-ideas.xlsx")
M2 = os.path.join(AE, "model-other-niches", "model-other-niches-ideas.xlsx")
M3 = os.path.join(AE, "study-trends", "study-trends-ideas.csv")
OUT = os.path.join(AE, "clubbed", "clubbed-ideas.csv")

COLS = ["Source method",
        "Brand fit", "Asset", "Format", "Distinct angle", "Total domains",
        "# comps", "# posts", "Beatability", "Effort", "Proof URLs",
        "RAG candidates",            # links + scores (Stage 2)
        "Topic pages we own",        # deterministic keyword catalogue links (Stage 2)
        "Reference links",           # de-duped union of the two link sets (Stage 2)
        "Reuse verdict", "Chosen links", "Why"]
# NOTE: links only — no "Candidate N content" columns. Page text is fetched on demand
# (look up the URL in content-database.csv, or web-fetch) by the Stage-3 judge / content creation.

def read_xlsx(path, sheet):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h).strip() if h is not None else "" for h in rows[0]]
    return [dict(zip(hdr, r)) for r in rows[1:] if any(c is not None for c in r)]

def g(row, key):
    v = row.get(key, "")
    return "" if v is None else v

m1 = [{"Source method": "Competitor",
       "Brand fit": g(r, "Brand fit"), "Asset": g(r, "Asset"), "Format": g(r, "Format"),
       "Distinct angle": g(r, "Distinct angle"), "Total domains": g(r, "Total domains"),
       "# comps": g(r, "# comps"), "Beatability": g(r, "Beat."), "Effort": g(r, "Effort"),
       "Proof URLs": g(r, "Backing URLs")} for r in read_xlsx(M1, "Ideas")]

with open(M3, newline="", encoding="utf-8") as f:
    m3 = [{"Source method": "Reddit",
           "Brand fit": g(r, "Brand fit"), "Asset": g(r, "Asset title"),
           "Distinct angle": g(r, "What it'd be"), "# posts": g(r, "# posts")}
          for r in csv.DictReader(f)]

m2 = [{"Source method": "Other-niche",
       "Brand fit": g(r, "Brand fit"), "Asset": g(r, "Asset"), "Format": g(r, "Format"),
       "Distinct angle": g(r, "Distinct angle"), "Beatability": g(r, "Beatability"),
       "Effort": g(r, "Effort"), "Proof URLs": g(r, "Evidence URLs")} for r in read_xlsx(M2, "ideas")]

os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(COLS)
    for block in (m1, m3, m2):
        for r in block:
            w.writerow([r.get(c, "") for c in COLS])

print(f"M1={len(m1)}  M3={len(m3)}  M2={len(m2)}  TOTAL={len(m1)+len(m3)+len(m2)}")
print("wrote", OUT)
