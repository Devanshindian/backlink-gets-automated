#!/usr/bin/env python3
"""Step H — build the two deliverables:
1. competitor-study-ideas.xlsx — Tab1 format summary, Tab2 ranked ideas (coloured by brand fit)
2. competitor-formats.xlsx — the master, one row per kept page, with Idea #
"""
import csv, sys, os
csv.field_size_limit(sys.maxsize)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HERE=os.path.dirname(__file__); BASE=os.path.join(HERE,'..')
CELL_MAX=32000
import re
ILLEGAL=re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')
def cap(s):
    s=ILLEGAL.sub(' ', s or '')
    return s[:CELL_MAX-3]+'...' if len(s)>CELL_MAX else s

HEAD=Font(bold=True,color='FFFFFF'); HEADFILL=PatternFill('solid',fgColor='1D2130')
FITFILL={'CORE':PatternFill('solid',fgColor='D9EAD3'),'TRANSPLANT':PatternFill('solid',fgColor='CFE2F3'),
         'ADJACENT':PatternFill('solid',fgColor='FFF2CC')}
NOWFILL=PatternFill('solid',fgColor='FCE5CD')

def style_header(ws,ncol):
    for c in range(1,ncol+1):
        cell=ws.cell(1,c); cell.font=HEAD; cell.fill=HEADFILL
        cell.alignment=Alignment(vertical='center',wrap_text=True)
    ws.freeze_panes='A2'

# ---------- DELIVERABLE 1: ideas ----------
wb=Workbook()
# Tab 1 — format summary
ws=wb.active; ws.title='Format summary'
fs=list(csv.DictReader(open(os.path.join(HERE,'format-summary.csv'))))
cols=['Format','# pages','Total Domains','Avg Domains','# competitors']
ws.append(cols)
for r in fs: ws.append([r['Format'],int(r['# pages']),int(r['Total Domains']),float(r['Avg Domains']),int(r['# competitors'])])
style_header(ws,len(cols))
for w,c in zip([34,10,14,12,14],'ABCDE'): ws.column_dimensions[c].width=w

# Tab 2 — ideas (ranked)
ws2=wb.create_sheet('Ideas')
ideas=list(csv.DictReader(open(os.path.join(HERE,'ideas.csv'),encoding='utf-8')))
cols2=['#','Build window','Brand fit','Asset','Our subject','Format','Total domains','# comps','# pages','Beat.','Effort','Distinct angle','Backing URLs']
ws2.append(cols2)
for o in ideas:
    ws2.append([int(o['IdeaNum']),o['BuildWindow'],o['BrandFit'],cap(o['Asset']),cap(o['OurSubject']),o['Format'],
                int(o['TotalDomains']),int(o['NumComps']),int(o['NumPages']),int(o['Beatability']),o['Effort'],
                cap(o['DistinctAngle']),cap(o['BackingURLs'])])
style_header(ws2,len(cols2))
for i,o in enumerate(ideas,start=2):
    fill=FITFILL.get(o['BrandFit'])
    if fill:
        for c in range(1,len(cols2)+1): ws2.cell(i,c).fill=fill
    if o['BuildWindow']=='NOW': ws2.cell(i,2).fill=NOWFILL; ws2.cell(i,2).font=Font(bold=True)
for w,c in zip([5,12,11,46,26,20,12,8,8,7,7,60,50],'ABCDEFGHIJKLM'): ws2.column_dimensions[c].width=w
for row in ws2.iter_rows(min_row=2):
    row[3].alignment=Alignment(wrap_text=True,vertical='top'); row[11].alignment=Alignment(wrap_text=True,vertical='top')
wb.save(os.path.join(BASE,'competitor-study-ideas.xlsx'))
print("wrote competitor-study-ideas.xlsx:",len(ideas),"ideas")

# ---------- DELIVERABLE 2: master ----------
wb2=Workbook(); ws=wb2.active; ws.title='Master'
rows=list(csv.DictReader(open(os.path.join(BASE,'competitor-formats.csv'),encoding='utf-8')))
cols=['Competitor','URL','Format','Domains','Backlinks','What it is','Asset','Brand fit','Angle gap','Distinct angle','Idea #','Notes']
ws.append(cols)
for r in rows:
    ws.append([r['Competitor'],r['URL'],r['Format'],int(r['Domains'] or 0),int(r['Backlinks'] or 0),
               cap(r['WhatItIs']),cap(r['Asset']),r['BrandFit'],cap(r['AngleGap']),cap(r['DistinctAngle']),
               r['IdeaNum'],cap(r['Notes'])])
style_header(ws,len(cols))
for i,r in enumerate(rows,start=2):
    fill=FITFILL.get(r['BrandFit'])
    if fill: ws.cell(i,8).fill=fill
for w,c in zip([13,50,22,9,10,60,40,10,45,40,7,30],'ABCDEFGHIJKL'): ws.column_dimensions[c].width=w
wb2.save(os.path.join(BASE,'competitor-formats.xlsx'))
print("wrote competitor-formats.xlsx:",len(rows),"rows")
