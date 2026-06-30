#!/usr/bin/env python3
"""Method 2 (Model Other Niches) — Testlify. Re-run from scratch against the updated recipe:
no Competitor coverage (overlap check is a merge-step concern), no Unfair advantage column
(folded into Distinct angle). 12-column shared schema. Only Step C (Ownability + Linkability)
drops ideas. Run: python3 build_ideas.py"""
import csv, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

OUT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
COLS = ["Idea #","Method","Brand fit","Asset","Format","Our topic","Distinct angle",
        "Source niche","Evidence URLs","Beatability","Effort","Notes"]
M = "M2 — Other Niches"

# Step B substance (Our topic, Distinct angle [names what we own], Asset) + Step D scores.
# One adaptation per swipe-row format. Controversy/opinion dropped at Step C (Linkability 2/4).
ideas = [
 dict(bf="CORE", fmt="Calculator", topic="Cost of a bad hire",
   asset="Bad-Hire Cost Calculator (per-role, with a sourced formula)",
   angle="Live, per-role calculator with a transparent sourced formula; uses our 4,500+ role taxonomy to pre-fill realistic salary/ramp defaults per role. No assessment company has a real calculator.",
   niche="finance / B2B SaaS", urls="nerdwallet.com/mortgage-calculator; hubspot.com/roi-calculator",
   beat=3, eff="S", notes="Headline: 'What does a bad hire actually cost you? [Free Calculator]'. 4-input form -> cost + formula + sharable PDF."),
 dict(bf="CORE", fmt="Infographic / data-visual", topic="Fair hiring process",
   asset="Anatomy of a Bias-Free Hiring Process (data-visual)",
   angle="One embeddable visual of a defensible process showing where bias enters and how to remove it, built on our fairness/process authority; embeds carry a credit link.",
   niche="various viral infographics", urls="informationisbeautiful.net; visualcapitalist.com",
   beat=3, eff="S", notes="Headline: 'Where bias sneaks into hiring (and how to stop it)'."),
 dict(bf="CORE", fmt="Statistics roundup", topic="Skills-based hiring statistics",
   asset="Skills-Based Hiring Statistics (sourced, updated yearly)",
   angle="Curated, every stat sourced, plus one original Testlify platform stat per section that no roundup can match.",
   niche="cybersecurity / marketing roundups", urls="varonis.com/blog/cybersecurity-statistics; hubspot.com/marketing-statistics",
   beat=3, eff="S", notes="One original platform stat per section is the moat."),
 dict(bf="CORE", fmt="Facts listicle", topic="Hiring & bias facts",
   asset="Surprising Facts About Hiring Bias (curated, sourced)",
   angle="Curiosity-driven facts list, each sourced, with a Testlify bias/assessment angle competitors' generic facts pages lack.",
   niche="general viral facts", urls="factretriever.com; mentalfloss.com",
   beat=3, eff="S", notes="Headline: '[#] facts about hiring bias most recruiters get wrong'."),
 dict(bf="CORE", fmt="Checklist", topic="Structured interviewing",
   asset="The Structured Interview Checklist (interactive + downloadable)",
   angle="Interactive checklist plus a scorecard download, built from our structured-interview product and scorecard templates.",
   niche="compliance (GDPR checklist)", urls="gdpr.eu/checklist; checklist.com",
   beat=3, eff="S", notes="Fast build; pairs with the JD generator."),
 dict(bf="CORE", fmt="Interactive explainer", topic="How AI screening works",
   asset="How AI Interview Scoring & Proctoring Works (interactive explainer)",
   angle="Interactive/animated explainer; we own the actual AI-scoring + proctoring product so we can show it accurately. Visual explainers get embedded + screenshotted.",
   niche="engineering / science (Animagraffs)", urls="animagraffs.com/jet-engine",
   beat=2, eff="M", notes="Headline: 'How AI Actually Scores an Interview'."),
 dict(bf="CORE", fmt="Benchmark / grader", topic="Hiring process fairness/quality",
   asset="How Fair Is Your Hiring? (free grader -> score + report)",
   angle="~12 questions -> a fairness/quality score + personalised report, built on our 'fair, defensible hiring' framework; the HubSpot-Grader engine, none in-niche. Shareable score drives embeds.",
   niche="marketing (HubSpot Website Grader)", urls="website.grader.com",
   beat=2, eff="M", notes="Headline: 'How fair is your hiring? Get your score'."),
 dict(bf="CORE", fmt="Quiz / typology", topic="Bias & fairness in hiring",
   asset="What's Your Hiring-Bias Profile? (quiz + shareable result)",
   angle="A B2B-appropriate quiz naming a recruiter's dominant biases + how to counter each, backed by our bias/fairness framework; shareable result page. No B2B hiring quiz exists in-niche.",
   niche="consumer psychology (16Personalities)", urls="16personalities.com",
   beat=2, eff="M", notes="B2B-framed (recruiter self-awareness), not consumer fluff."),
 dict(bf="CORE", fmt="Generator", topic="Job descriptions / role definitions",
   asset="Skills-First Job Description Generator (bias-checked, test-linked)",
   angle="Outputs a skills-first JD with a bias-language check and a linked test per skill, drawn from our 4,500+ role taxonomy and test catalog.",
   niche="SEO / branding (name generators)", urls="shopify.com/tools/business-name-generator; semrush.com/generate",
   beat=3, eff="M", notes="Pairs with the structured-interview checklist."),
 dict(bf="CORE", fmt="Complete-list listicle", topic="Skills to test per role",
   asset="The Complete List of Skills to Test, by Role (500+)",
   angle="Exhaustive role -> skills -> test mapping, a byproduct of our 3,500+ tests across 4,500+ roles; each skill links to an actual test. Nobody owns the canonical list.",
   niche="SEO (Google 200 Ranking Factors)", urls="backlinko.com/google-ranking-factors",
   beat=3, eff="M", notes="Strong product tie-in; hard for a non-assessment company to match."),
 dict(bf="CORE", fmt="Definitive guide + template", topic="Skills-based hiring methodology",
   asset="Skills-Based Hiring: The Definitive Guide + Scorecard Template",
   angle="The canonical guide plus a free downloadable scorecard, anchored in our assessment product and fairness authority.",
   niche="SEO (On-Page SEO Definitive Guide)", urls="backlinko.com/on-page-seo",
   beat=3, eff="M", notes="Overlaps Method 1's guide/template lanes -> the merge will dedupe."),
 dict(bf="CORE", fmt="Glossary / dictionary", topic="Assessment & hiring terminology",
   asset="Hiring & Assessment Glossary with Bundled Micro-Tests",
   angle="Every term bundled with a 3-question micro-test or template, drawn from our test catalog; competitor glossaries are definition-only.",
   niche="niche glossaries (poker/golf terms)", urls="merriam-webster.com",
   beat=3, eff="M", notes="Overlaps Method 1's strong glossary lane -> the merge will dedupe."),
 dict(bf="CORE", fmt="Template / asset library", topic="Hiring templates",
   asset="The Hiring Template Library (scorecards, JDs, rubrics)",
   angle="A hub of free downloadable hiring templates (scorecards, JD frameworks, interview rubrics) generated from our role taxonomy and structured-interview product.",
   niche="marketing (HubSpot Template Library)", urls="hubspot.com/resources/templates",
   beat=3, eff="M", notes="Overlaps Method 1's strong template lane -> the merge will dedupe."),
 dict(bf="CORE", fmt="Free practice test / question bank", topic="Practice tests / sample questions",
   asset="Free Practice Tests + Public Sample Question Bank",
   angle="High-volume free practice tests with public sample questions per role, generated from our 3,500+ test library; competitor test pages paywall the questions.",
   niche="test-prep (LeetCode; SAT prep)", urls="leetcode.com; khanacademy.org/sat",
   beat=3, eff="M", notes="Overlaps Method 1's strong free-test lane -> the merge will dedupe."),
 dict(bf="CORE", fmt="Case study (results)", topic="Hiring outcomes",
   asset="How [Customer] Cut Time-to-Hire with Skills Tests (results case study)",
   angle="A concrete, numbers-led customer result (time-to-hire, quality-of-hire) only we can publish from our own customer base.",
   niche="SEO/marketing case studies", urls="backlinko.com/seo-case-study",
   beat=2, eff="M", notes="Needs a customer willing to share real numbers."),
 dict(bf="CORE", fmt="Original research", topic="Predictive validity of skills tests vs resumes",
   asset="We Analyzed [N] Assessments: What Actually Predicts a Good Hire",
   angle="A large-scale look from real assessment outcomes; competitors' data reports are opinion surveys, not platform data. Original research is the most-cited format.",
   niche="SEO / SaaS (Backlinko)", urls="backlinko.com/search-engine-ranking-factors",
   beat=2, eff="L", notes="NEEDS first-party assessment data release. Highest ceiling if freed."),
 dict(bf="CORE", fmt="State-of survey report", topic="Skills-based hiring adoption",
   asset="The State of Skills-Based Hiring (annual survey report)",
   angle="A recurring, branded annual combining a survey of our 1,500+ teams with platform data; competitor reports are one-off and unbranded.",
   niche="dev / marketing (State of DevOps; SO Survey)", urls="stackoverflow.co/dev-survey; puppet.com/state-of-devops",
   beat=2, eff="L", notes="Recurring; needs a survey run. Becomes the canonical annual citation."),
 dict(bf="CORE", fmt="Award + badge", topic="Candidate experience",
   asset="Best Companies for Candidate Experience (annual award + badge)",
   angle="Annual award; winners embed a badge that links back (a recurring ego-bait link engine), judged on candidate-experience signals from our completion/satisfaction metrics + authority.",
   niche="HR / employer brand (Best Places to Work)", urls="greatplacetowork.com/best-workplaces",
   beat=2, eff="L", notes="Recurring; badge backlinks compound yearly. Needs program ops."),
 dict(bf="CORE", fmt="Branded index", topic="Cross-industry skills gaps",
   asset="The Skills-Gap Index (which skills each industry is short on, annual)",
   angle="A named, recurring index ranking skill shortages by industry/role from our aggregate test pass-rate data; journalists cite indices by name. No assessment vendor publishes one.",
   niche="economics (Big Mac Index)", urls="economist.com/big-mac-index",
   beat=1, eff="L", notes="Recurring; NEEDS first-party data release. D2-validate before building."),
 dict(bf="CORE", fmt="Rankings", topic="In-demand skills trends",
   asset="Most In-Demand Skills, Ranked (quarterly)",
   angle="Quarterly ranking of skills by test demand/growth from our platform data; a recurring citation hook that updates each cycle.",
   niche="education / sports rankings", urls="usnews.com/best-colleges",
   beat=1, eff="L", notes="Recurring; NEEDS first-party data release."),
 dict(bf="CORE", fmt="Map / geo viz", topic="Regional skills demand / remote hiring",
   asset="Remote Skills-Demand Map (by country/region)",
   angle="Interactive map of skill demand by region from our aggregate platform geo data; geo data-viz is a reliable embed magnet.",
   niche="geo data viz (cost-of-living maps)", urls="visualcapitalist.com",
   beat=1, eff="L", notes="NEEDS first-party data release."),
 dict(bf="TRANSPLANT", fmt="Salary benchmark", topic="Value of demonstrated skills (transplanted from compensation)",
   asset="Pay-by-Skill Benchmark (what proven skills are worth)",
   angle="Instead of generic salary data, benchmark pay against demonstrated skill level using our skill-proficiency data; pulls salary-tool link demand into our skills lens.",
   niche="compensation (Levels.fyi; PayScale)", urls="levels.fyi; payscale.com",
   beat=1, eff="L", notes="transplanted from: compensation. NEEDS first-party data release."),
]

dropped = [
 ("Controversy / opinion ('Why resumes predict worse')","C — Linkability 2/4",
  "No citable number; opinion pieces pull shares, not links. Fails the >=3/4 Linkability gate."),
]

# Step E sort: Brand fit (CORE->TRANSPLANT->ADJACENT) -> Evidence x Beatability desc -> Effort S first.
bf_rank={"CORE":0,"TRANSPLANT":1,"ADJACENT":2}; eff_rank={"S":0,"M":1,"L":2}
def evidence(u): return len([x for x in u.split(";") if x.strip()])
ideas.sort(key=lambda r:(bf_rank[r["bf"]], -(evidence(r["urls"])*r["beat"]), eff_rank[r["eff"]]))

rows=[[f"M2-{i:03d}",M,r["bf"],r["asset"],r["fmt"],r["topic"],r["angle"],r["niche"],
       r["urls"],r["beat"],r["eff"],r["notes"]] for i,r in enumerate(ideas,1)]

with open(os.path.join(OUT,"model-other-niches-ideas.csv"),"w",newline="") as f:
    w=csv.writer(f); w.writerow(COLS); w.writerows(rows)

wb=openpyxl.Workbook(); ws=wb.active; ws.title="ideas"
fills={"CORE":PatternFill("solid",fgColor="C6EFCE"),
       "TRANSPLANT":PatternFill("solid",fgColor="BDD7EE"),
       "ADJACENT":PatternFill("solid",fgColor="FFE699")}
ws.append(COLS)
for c in ws[1]: c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="1D2130")
for row in rows:
    ws.append(row)
    for c in ws[ws.max_row]: c.fill=fills.get(row[2],PatternFill())
for i,wd in enumerate([8,16,11,42,22,26,60,28,34,7,7,52],1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=wd
for r in ws.iter_rows():
    for c in r: c.alignment=Alignment(vertical="top",wrap_text=True)
ws.freeze_panes="A2"
wb.save(os.path.join(OUT,"model-other-niches-ideas.xlsx"))

print(f"KEPT ideas: {len(rows)} (12-col schema, no coverage/advantage)")
print(f"  CORE: {sum(1 for r in rows if r[2]=='CORE')}  TRANSPLANT: {sum(1 for r in rows if r[2]=='TRANSPLANT')}")
print(f"DROPPED at Step C: {len(dropped)} -> {dropped[0][0]}")
