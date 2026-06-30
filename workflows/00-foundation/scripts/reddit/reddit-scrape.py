#!/usr/bin/env python3
"""
reddit-scrape.py — free Reddit trend-scraper for the Asset Engine (Method 4: Study Trends).

Scrapes Reddit's PUBLIC old.reddit HTML (no API key, no login). Four verbs:

  discover "<topic>" [--limit 15]
      Find subreddits for a topic + their subscriber counts + descriptions.
  posts <subreddit> [--sort top] [--time year] [--limit 25]
      Top/hot/new/rising posts for one subreddit (title, score, comments, flair, date, url).
  comments <post-url> [--threads 15]
      Full nested comment tree for one post (top N top-level threads, all sub-replies).
  run <sub1,sub2,...> [--posts 25] [--threads 15] [--sort top] [--time year] --out <file.xlsx>
      The pipeline: for each subreddit -> top posts -> each post's comment tree,
      written to ONE .xlsx (tabs: Subreddits / Posts / Comments) + raw/ JSON backup.

Access ladder: this is the FREE default. If old.reddit gets blocked, the only fallback is the official
Reddit API (PRAW) with an approved account.
Honest limits: Reddit exposes NO view counts — we rank by upvote score + comment count only.
"""
import sys, re, json, time, html, argparse, os, subprocess, tempfile
import urllib.request, urllib.parse, urllib.error
from html.parser import HTMLParser

UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36"
THROTTLE = 2.5  # seconds between requests, to avoid 429s
OCR_BIN = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ocr-mac")  # native macOS OCR (optional)


def ocr_image(image_url, referer=""):
    """Download an image post's image and OCR it with the native macOS tool. '' if OCR unavailable/fails."""
    if not image_url or not os.path.exists(OCR_BIN):
        return ""
    try:
        ext = os.path.splitext(image_url)[1][:5] or ".jpg"
        fd, tmp = tempfile.mkstemp(suffix=ext)
        os.close(fd)
        req = urllib.request.Request(image_url, headers={"User-Agent": UA, "Referer": referer})
        with urllib.request.urlopen(req, timeout=30) as r, open(tmp, "wb") as f:
            f.write(r.read())
        out = subprocess.run([OCR_BIN, tmp], capture_output=True, text=True, timeout=60)
        os.remove(tmp)
        time.sleep(1.0)
        return out.stdout.strip()
    except Exception:
        return ""


def fetch(url, tries=5):
    """GET a URL with a browser UA + throttle + backoff on 429 and network drops."""
    for attempt in range(tries):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": UA})
            with urllib.request.urlopen(req, timeout=30) as r:
                data = r.read().decode("utf-8", "replace")
            time.sleep(THROTTLE)
            return data
        except urllib.error.HTTPError as e:
            if e.code == 429:
                time.sleep(8 * (attempt + 1))  # back off harder on rate-limit
                continue
            raise
        except urllib.error.URLError:
            time.sleep(5 * (attempt + 1))  # network/DNS hiccup (e.g. laptop wake) — wait + retry
            continue
    raise RuntimeError(f"failed after {tries} tries: {url}")


def clean(t):
    return html.unescape(re.sub(r"\s+", " ", t or "")).strip()


# ---------- POSTS ----------
def parse_posts(page):
    """Parse an old.reddit listing page into post dicts."""
    rows = []
    for block in re.split(r'(?=<div [^>]*class="[^"]*\bthing\b)', page):
        if 'data-fullname="t3_' not in block:
            continue
        def attr(n):
            m = re.search(r'data-%s="([^"]*)"' % n, block)
            return m.group(1) if m else ""
        title = re.search(r'class="title may-blank[^"]*"[^>]*>([^<]+)<', block)
        flair = re.search(r'class="linkflairlabel[^"]*"[^>]*title="([^"]*)"', block)
        if not attr("score"):
            continue
        linkurl = attr("url")  # data-url: the image/external link for link posts
        rows.append({
            "post_id": attr("fullname")[3:],
            "subreddit": attr("subreddit"),
            "score": int(attr("score") or 0),
            "comment_count": int(attr("comments-count") or 0),
            "flair": clean(flair.group(1)) if flair else "",
            "date": attr("timestamp"),
            "author": attr("author"),
            "title": clean(title.group(1)) if title else "",
            "image_url": linkurl if re.search(r"(i\.redd\.it|i\.imgur\.com|preview\.redd\.it)", linkurl) else "",
            "url": "https://old.reddit.com" + attr("permalink"),
        })
    return rows


# ---------- COMMENTS (nested tree) ----------
class CommentTreeParser(HTMLParser):
    """Walks old.reddit comment HTML, tracking div nesting to recover depth + parent."""
    def __init__(self):
        super().__init__()
        self.divcount = 0
        self.stack = []          # open comment frames (ancestry)
        self.comments = []
        self.in_md_at = None     # divcount where the current comment's body opened
        self.cur_text = None

    def handle_starttag(self, tag, attrs):
        a = dict(attrs)
        if tag == "div":
            self.divcount += 1
            cls = a.get("class", "")
            fn = a.get("data-fullname", "")
            if "thing" in cls and "comment" in cls and fn.startswith("t1_"):
                c = {
                    "comment_id": fn[3:],
                    "parent_id": self.stack[-1]["comment_id"] if self.stack else "",
                    "depth": len(self.stack),
                    "opendiv": self.divcount,
                    "subreddit": a.get("data-subreddit", ""),
                    "author": a.get("data-author", ""),
                    "score": None,
                    "text": "",
                }
                self.stack.append(c)
                self.comments.append(c)
            elif cls.strip().startswith("md") and self.stack and self.in_md_at is None:
                # body of the current (deepest open) comment
                self.in_md_at = self.divcount
                self.cur_text = []
        elif tag == "span":
            # score lives in <span class="score unvoted" title="N points">
            if a.get("class") == "score unvoted" and self.stack and self.stack[-1]["score"] is None:
                m = re.match(r"(-?\d+)", a.get("title", ""))
                if m:
                    self.stack[-1]["score"] = int(m.group(1))

    def handle_data(self, data):
        if self.in_md_at is not None and self.cur_text is not None:
            self.cur_text.append(data)

    def handle_endtag(self, tag):
        if tag == "div":
            if self.in_md_at is not None and self.divcount == self.in_md_at:
                if self.stack:
                    self.stack[-1]["text"] = clean("".join(self.cur_text))
                self.in_md_at = None
                self.cur_text = None
            if self.stack and self.stack[-1]["opendiv"] == self.divcount:
                self.stack.pop()
            self.divcount -= 1


def parse_post_body(page):
    """Extract the post's own body text (selftext) from its page — '' for image/link posts."""
    i = page.find('data-fullname="t3_')
    if i < 0:
        return ""
    block = page[i:page.find("commentarea", i)]
    s = block.find('<div class="md">')
    if s < 0:
        return ""
    sub = block[s + len('<div class="md">'):]
    end = sub.find("</form>")               # selftext md sits inside the post's usertext <form>
    body = sub[:end] if end >= 0 else sub[:4000]
    return clean(re.sub(r"<[^>]+>", " ", body))


def parse_comments(page, max_threads=15):
    p = CommentTreeParser()
    p.feed(page)
    rows = [c for c in p.comments if c["score"] is not None or c["text"]]
    # keep only the first `max_threads` top-level threads + all their descendants
    kept, threads = [], 0
    keeping = False
    for c in rows:
        if c["depth"] == 0:
            threads += 1
            keeping = threads <= max_threads
        if keeping:
            c.pop("opendiv", None)
            kept.append(c)
    return kept


# ---------- VERBS ----------
def discover(topic, limit=15):
    page = fetch(f"https://old.reddit.com/subreddits/search?q={urllib.parse.quote(topic)}")
    out = []
    for m in re.finditer(r'<a[^>]*href="(?:https?://old\.reddit\.com)?/r/([^/"]+)/?"[^>]*class="[^"]*\btitle\b[^"]*"[^>]*>([^<]+)</a>', page):
        name = m.group(1)
        desc = clean(m.group(2))
        if name.lower() in {x["subreddit"].lower() for x in out}:
            continue
        out.append({"subreddit": name, "subscribers": None, "description": desc})
        if len(out) >= limit:
            break
    # NOTE: old.reddit no longer exposes subscriber counts in the HTML (2026-06), so we leave
    # subscribers=None here. Don't gate on member size — vet candidates by activity (top-post
    # upvotes/comments via the `posts` command) before the approval gate.
    return out


def get_posts(sub, sort="top", t="year", limit=25):
    url = f"https://old.reddit.com/r/{sub}/{sort}/?t={t}&limit={min(limit,100)}"
    return parse_posts(fetch(url))[:limit]


def get_comments(post_url, threads=15):
    """Fetch a post page once; return (post_body, comment_tree)."""
    url = post_url.split("?")[0].rstrip("/") + "/?sort=top&limit=500"
    page = fetch(url)
    return parse_post_body(page), parse_comments(page, max_threads=threads)


def cmd_run(subs, posts_n, threads, sort, t, out_path, raw_dir):
    os.makedirs(raw_dir, exist_ok=True)
    all_posts, all_comments, sub_rows = [], [], []
    for sub in subs:
        pf = f"{raw_dir}/{sub}-posts.json"
        if os.path.exists(pf):                      # resume: reuse saved post list
            posts = json.load(open(pf))
            print(f"[{sub}] posts (cached: {len(posts)})", file=sys.stderr)
        else:
            print(f"[{sub}] posts...", file=sys.stderr)
            posts = get_posts(sub, sort, t, posts_n)
            json.dump(posts, open(pf, "w"), ensure_ascii=False)
        sub_rows.append({"subreddit": sub, "posts_pulled": len(posts)})
        all_posts += posts
        for i, p in enumerate(posts, 1):
            cf = f"{raw_dir}/{p['post_id']}-comments.json"
            if os.path.exists(cf):                  # resume: skip already-scraped comments
                all_comments += json.load(open(cf))
                print(f"  [{sub}] comments {i}/{len(posts)} (cached)", file=sys.stderr)
                continue
            print(f"  [{sub}] comments {i}/{len(posts)}", file=sys.stderr)
            try:
                body, cmts = get_comments(p["url"], threads)
                p["body"] = body
                p["image_text"] = ocr_image(p.get("image_url", ""), p["url"])  # OCR image posts (if ocr-mac present)
                for c in cmts:
                    c["post_id"] = p["post_id"]
                    c["subreddit"] = c["subreddit"] or sub
                json.dump(cmts, open(cf, "w"), ensure_ascii=False)
                all_comments += cmts
            except Exception as e:
                print(f"    skip {p['post_id']}: {e}", file=sys.stderr)
        json.dump(posts, open(pf, "w"), ensure_ascii=False)   # re-save with bodies attached
    write_xlsx(out_path, sub_rows, all_posts, all_comments)
    print(f"\nDONE -> {out_path}\n  {len(all_posts)} posts, {len(all_comments)} comments across {len(subs)} subreddits", file=sys.stderr)


def cmd_build(raw_dir, out_path):
    """Assemble the .xlsx from whatever raw/ JSON exists (works on partial runs)."""
    import glob
    sub_rows, all_posts, all_comments = [], [], []
    for pf in sorted(glob.glob(f"{raw_dir}/*-posts.json")):
        sub = os.path.basename(pf)[:-len("-posts.json")]
        posts = json.load(open(pf))
        sub_rows.append({"subreddit": sub, "posts_pulled": len(posts)})
        all_posts += posts
    for cf in sorted(glob.glob(f"{raw_dir}/*-comments.json")):
        all_comments += json.load(open(cf))
    write_xlsx(out_path, sub_rows, all_posts, all_comments)
    print(f"built {out_path}: {len(all_posts)} posts, {len(all_comments)} comments", file=sys.stderr)


def top_comments_by_post(comments, n=5):
    """For each post_id, the n highest-scored TOP-LEVEL comments, formatted inline for the Posts tab."""
    by_post = {}
    for c in comments:
        txt = (c.get("text") or "").strip()
        if c.get("depth") == 0 and txt and txt not in ("[removed]", "[deleted]"):
            by_post.setdefault(c["post_id"], []).append(c)
    out = {}
    for pid, lst in by_post.items():
        lst.sort(key=lambda x: -(x.get("score") or 0))
        out[pid] = "\n".join(f"[▲{c.get('score')}] {c.get('text','')}" for c in lst[:n])
    return out


def write_xlsx(path, subs, posts, comments):
    import openpyxl
    wb = openpyxl.Workbook()
    s1 = wb.active; s1.title = "Subreddits"
    s1.append(["subreddit", "posts_pulled"])
    for r in subs: s1.append([r["subreddit"], r["posts_pulled"]])
    s2 = wb.create_sheet("Posts")
    topc = top_comments_by_post(comments, n=15)  # top 15 comments per post, inline
    # comment_count sits right before top_comments: reads "[# comments] -> [the 15 top comments]"
    pcols = ["post_id", "subreddit", "score", "flair", "date", "author", "title", "body",
             "image_url", "image_text", "comment_count", "top_comments", "url"]
    s2.append(pcols)
    for r in sorted(posts, key=lambda x: -x["score"]):
        r = {**r, "top_comments": topc.get(r["post_id"], "")}
        s2.append([r.get(c, "") for c in pcols])
    s3 = wb.create_sheet("Comments")
    # `score` = this comment's upvotes (the ▲N number); top_comments above is sorted by it, top 15 per post
    ccols = ["post_id", "subreddit", "comment_id", "parent_id", "depth", "score", "author", "text"]
    s3.append(ccols)
    for r in comments:
        s3.append([r.get(c, "") for c in ccols])
    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest="cmd", required=True)
    d = sub.add_parser("discover"); d.add_argument("topic"); d.add_argument("--limit", type=int, default=15)
    p = sub.add_parser("posts"); p.add_argument("subreddit"); p.add_argument("--sort", default="top"); p.add_argument("--time", default="year"); p.add_argument("--limit", type=int, default=25)
    c = sub.add_parser("comments"); c.add_argument("url"); c.add_argument("--threads", type=int, default=15)
    r = sub.add_parser("run"); r.add_argument("subs"); r.add_argument("--posts", type=int, default=25); r.add_argument("--threads", type=int, default=15); r.add_argument("--sort", default="top"); r.add_argument("--time", default="year"); r.add_argument("--out", required=True); r.add_argument("--raw", default="raw")
    b = sub.add_parser("build"); b.add_argument("--raw", default="raw"); b.add_argument("--out", required=True)
    args = ap.parse_args()
    if args.cmd == "discover":
        print(json.dumps(discover(args.topic, args.limit), ensure_ascii=False, indent=2))
    elif args.cmd == "posts":
        print(json.dumps(get_posts(args.subreddit, args.sort, args.time, args.limit), ensure_ascii=False, indent=2))
    elif args.cmd == "comments":
        body, cmts = get_comments(args.url, args.threads)
        print(json.dumps({"body": body, "comments": cmts}, ensure_ascii=False, indent=2))
    elif args.cmd == "run":
        cmd_run([s.strip() for s in args.subs.split(",") if s.strip()], args.posts, args.threads, args.sort, args.time, args.out, args.raw)
    elif args.cmd == "build":
        cmd_build(args.raw, args.out)


if __name__ == "__main__":
    import urllib.parse, urllib.error
    main()
